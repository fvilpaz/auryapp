"""
Importa todos los turnos del Excel data/turnos.xlsx a la BD SQLite.
Lee el color de cada celda para distinguir tipos de día libre.
Borra primero los turnos existentes en el rango del Excel y reimporta todo.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import sqlite3
from datetime import datetime, date, time

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'db.sqlite3')
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'turnos.xlsx')

EMPLOYEE_MAP = {
    'AURY':     1,
    'EDU':      2,
    'ANTONIO':  3,
    'EVA':      4,
    'INMA':     5,
    'LAURA':    6,
    'VANE':     7,
    'ARLEN':    8,
    'SERGIO':   9,
    'JHONATAN': 10,
    'JESSICA':  11,
    'PAULA':    12,
    'LAURA M.': 13,
}

# Color de celda → estado (leído de la leyenda del Excel)
COLOR_ESTADO_MAP = {
    'FFFF9300': 'inamovible',        # naranja  → inamovible
    'FF7030A0': 'libre',             # morado   → día libre
    'FFFF00AE': 'finde_largo',       # rosa/mag → finde largo
    'FFFFC000': 'vacaciones',        # ámbar    → vacaciones
    'FFFF0000': 'baja',              # rojo     → baja
    # color de tema (error al leer) → libre_vacaciones
}


def get_cell_estado(cell):
    """Lee el color de la celda L y devuelve el estado Django correspondiente."""
    try:
        rgb = str(cell.fill.fgColor.rgb)
        return COLOR_ESTADO_MAP.get(rgb, 'libre')
    except Exception:
        # Color de tema (no RGB directo) → libre_vacaciones
        return 'libre_vacaciones'


def parse_time(s):
    s = str(s).strip()
    if len(s) <= 2:
        # "12" → 12:00, no "00:12"
        try:
            return time(int(s), 0)
        except ValueError:
            return None
    s = s.zfill(4)
    try:
        return time(int(s[:2]), int(s[2:]))
    except ValueError:
        return None


def parse_shift(entry, cell):
    """Devuelve (estado, hora_inicio, hora_fin) o None si no hay dato."""
    if entry is None:
        return None
    v = str(entry).strip()
    if not v:
        return None
    u = v.upper()

    if u == 'L':
        estado = get_cell_estado(cell)
        return (estado, None, None)

    if u == 'VACACIONES':
        return ('vacaciones', None, None)
    if u in ('B', 'BAJA'):
        return ('baja', None, None)
    if u in ('EVENTO', 'MONTAJE', 'MONTAJE '):
        return ('trabajo', None, None)
    if u in ('RECUPERA', 'RECUPERA '):
        return ('libre', None, None)

    if '-' in v:
        parts = v.replace(' ', '').split('-')
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            h_ini = parse_time(parts[0])
            h_fin = parse_time(parts[1])
            if h_ini:
                return ('trabajo', h_ini, h_fin)
    return None


def compute_dates(ref_date, day_numbers):
    month, year = ref_date.month, ref_date.year
    dates = []
    prev = 0
    for d in day_numbers:
        if not isinstance(d, int):
            dates.append(None)
            continue
        if d < prev and prev >= 20:
            month += 1
            if month > 12:
                month, year = 1, year + 1
        dates.append(date(year, month, d))
        prev = d
    return dates


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    # Necesitamos las celdas con fill, no solo values_only
    all_rows_cells = list(ws.iter_rows())
    all_rows_vals = list(ws.iter_rows(values_only=True))

    shifts = []
    i = 0
    while i < len(all_rows_vals):
        row_v = all_rows_vals[i]
        row_c = all_rows_cells[i]

        if not isinstance(row_v[0], datetime):
            i += 1
            continue

        ref = row_v[0].date()
        dates = compute_dates(ref, list(row_v[1:8]))
        i += 1

        while i < len(all_rows_vals):
            r_v = all_rows_vals[i]
            r_c = all_rows_cells[i]

            if isinstance(r_v[0], datetime):
                break

            name_raw = r_v[0]
            if name_raw is None:
                i += 1
                continue

            name = str(name_raw).strip().upper()
            if name == 'PERSONAL':
                i += 1
                continue

            emp_id = EMPLOYEE_MAP.get(name)
            if emp_id is None:
                i += 1
                continue

            day_entries = list(r_v[1:8])
            day_cells   = list(r_c[1:8])

            # Fila de horas justo debajo
            hours = [None] * 7
            if i + 1 < len(all_rows_vals) and all_rows_vals[i + 1][0] is None:
                hours = [all_rows_vals[i + 1][k] for k in range(1, 8)]

            # Semana de vacaciones completa
            is_vac = any(
                str(e).strip().upper() == 'VACACIONES'
                for e in day_entries if e is not None
            )

            for k, (d, entry, cell) in enumerate(zip(dates, day_entries, day_cells)):
                if d is None:
                    continue
                if is_vac:
                    shifts.append((emp_id, d.isoformat(), 'vacaciones', None, None, 0))
                    continue
                result = parse_shift(entry, cell)
                if result is None:
                    continue
                estado, h_ini, h_fin = result
                horas = hours[k] if isinstance(hours[k], int) else 0
                h_ini_str = h_ini.strftime('%H:%M:%S') if h_ini else None
                h_fin_str = h_fin.strftime('%H:%M:%S') if h_fin else None
                shifts.append((emp_id, d.isoformat(), estado, h_ini_str, h_fin_str, horas))

            i += 1

    return shifts


def main():
    shifts = parse_excel()
    if not shifts:
        print("No se encontraron turnos en el Excel.")
        return

    dates = [s[1] for s in shifts]
    date_min, date_max = min(dates), max(dates)
    print(f"Turnos parseados: {len(shifts)}  |  rango: {date_min} → {date_max}")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM personal_turno WHERE fecha >= ? AND fecha <= ?",
        (date_min, date_max)
    )
    print(f"Turnos eliminados del rango: {cur.rowcount}")

    cur.executemany(
        """INSERT INTO personal_turno
           (empleado_id, fecha, estado, hora_inicio, hora_fin, horas, espacio_id, color_override)
           VALUES (?, ?, ?, ?, ?, ?, NULL, '')""",
        shifts
    )
    conn.commit()
    print(f"Turnos insertados: {cur.rowcount}")

    cur.execute("""
        SELECT e.nombre, COUNT(*) as n
        FROM personal_turno t
        JOIN personal_empleado e ON t.empleado_id = e.id
        WHERE t.fecha >= ? AND t.fecha <= ?
        GROUP BY e.nombre ORDER BY e.nombre
    """, (date_min, date_max))
    print("\nTurnos por empleado:")
    for row in cur.fetchall():
        print(f"  {row[0]:<12} {row[1]}")

    # Resumen de estados
    cur.execute("""
        SELECT estado, COUNT(*) FROM personal_turno
        WHERE fecha >= ? AND fecha <= ?
        GROUP BY estado ORDER BY COUNT(*) DESC
    """, (date_min, date_max))
    print("\nPor estado:")
    for row in cur.fetchall():
        print(f"  {row[0]:<20} {row[1]}")

    conn.close()


if __name__ == '__main__':
    main()
